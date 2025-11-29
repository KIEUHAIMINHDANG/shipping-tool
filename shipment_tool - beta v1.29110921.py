import pandas as pd
import os
import glob
import re
import threading
import tkinter as tk
from tkinter import filedialog, scrolledtext, messagebox, ttk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import subprocess
import platform

# ==============================================================================
# PART 1: CORE LOGIC - SCANNING & PARSING
# ==============================================================================

def normalize_text(text):
    if pd.isna(text): return ""
    text = str(text).lower()
    text = re.sub(r'[\n\r\t]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def find_headers_robust(df):
    """Scans for 8 specific columns to build the complex string."""
    limit = min(30, len(df))
    for r in range(limit):
        row_clean = [normalize_text(val) for val in df.iloc[r]]
        idxs = {
            'name': -1, 'code': -1, 'qty_per_ctn': -1, 'range': -1, 
            'ctn_count': -1, 'total_qty': -1, 'nw': -1, 'gw': -1
        }
        for i, val in enumerate(row_clean):
            if "tên sản phẩm" in val or "item description" in val: idxs['name'] = i
            if "mã số" in val or "item code" in val: idxs['code'] = i
            if "1 thùng" in val or "pcs/ctn" in val or ("qty" in val and "/" in val): idxs['qty_per_ctn'] = i
            if "dải số thùng" in val or "carton no" in val: idxs['range'] = i
            if "số lượng carton" in val or "ctns" in val or "carton qty" in val: idxs['ctn_count'] = i
            if "lượng xuất" in val or "total qty" in val: idxs['total_qty'] = i
            
            if "total n.w" in val or "t.n.w" in val: idxs['nw'] = i
            elif "n.w" in val and idxs['nw'] == -1: idxs['nw'] = i
            
            if "total g.w" in val or "t.g.w" in val: idxs['gw'] = i
            elif "g.w" in val and idxs['gw'] == -1: idxs['gw'] = i

        if idxs['name'] != -1 and idxs['total_qty'] != -1:
            return r, idxs
    return None, None

def process_files(input_folder, log_callback, progress_callback, finish_callback):
    """Main scanning loop."""
    file_pattern = "Tuần *.xlsm" 
    ignored_sheet_names = ['mail', 'mau_mui_tuan', 'sheet1', 'sheet 1']
    search_path = os.path.join(input_folder, file_pattern)
    files = glob.glob(search_path)

    if not files:
        log_callback("ERROR: No 'Tuần *.xlsm' files found.")
        return

    total_files = len(files)
    master_data = []

    for f_idx, file_path in enumerate(files):
        filename = os.path.basename(file_path)
        week_match = re.search(r"Tuần\s*(\d+)", filename, re.IGNORECASE)
        week_num = int(week_match.group(1)) if week_match else 0
        
        log_callback(f"Scanning ({f_idx + 1}/{total_files}): {filename}...")
        progress_val = ((f_idx + 1) / total_files) * 100
        progress_callback(progress_val)

        try:
            xls = pd.ExcelFile(file_path, engine='openpyxl')
            for sheet_name in xls.sheet_names:
                if sheet_name.lower() in ignored_sheet_names: continue
                try:
                    df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=100)
                    header_row, idxs = find_headers_robust(df_raw)
                    if header_row is None: continue

                    pod_val = "N/A"
                    if len(df_raw) > 6:
                        val = df_raw.iloc[6, 2]
                        if pd.notna(val): pod_val = str(val).strip().replace('\n', ' ')

                    strip_items = []
                    for i in range(header_row + 1, len(df_raw)):
                        row = df_raw.iloc[i]
                        
                        item_name = row[idxs['name']]
                        if pd.isna(item_name): continue
                        str_name = str(item_name).strip().replace('\n', ' ')
                        
                        name_upper = str_name.upper()
                        if any(x in name_upper for x in ["TOTAL", "MÃ CÂN", "LÁI XE", "SIGNATURE", "GIÁ GỖ"]): break

                        # Extract fields
                        str_code = f"({str(row[idxs['code']]).strip()})" if idxs['code'] != -1 and pd.notna(row[idxs['code']]) else ""
                        
                        qty_per_str = ""
                        if idxs['qty_per_ctn'] != -1:
                            val = row[idxs['qty_per_ctn']]
                            if pd.notna(val) and isinstance(val, (int, float)): qty_per_str = f"QTY: {int(val):,} pcs"

                        ctn_count_str = ""
                        if idxs['ctn_count'] != -1:
                            val = row[idxs['ctn_count']]
                            if pd.notna(val) and isinstance(val, (int, float)): ctn_count_str = f"{int(val):,} cartons"

                        range_str = ""
                        if idxs['range'] != -1:
                            try:
                                base = idxs['range']
                                c1 = str(row[base]).strip().replace('.0', '') if pd.notna(row[base]) else ""
                                c2 = str(row[base+1]).strip() if pd.notna(row[base+1]) else ""
                                c3 = str(row[base+2]).strip().replace('.0', '') if pd.notna(row[base+2]) else ""
                                if c1 or c3: range_str = f"[Ctn: {c1}{c2}{c3}]"
                            except: pass

                        total_qty_str = ""
                        val = row[idxs['total_qty']]
                        try:
                             if isinstance(val, str): val = float(val.replace(',', '').replace('.', ''))
                        except: pass
                        if isinstance(val, (int, float)) and val > 0: total_qty_str = f"{int(val):,} pcs"

                        # NW/GW Split Logic
                        nw_str, gw_str = "", ""
                        is_combined = (idxs['nw'] != -1 and idxs['gw'] != -1 and idxs['nw'] == idxs['gw'])
                        if is_combined:
                            raw = str(row[idxs['nw']]).strip() if pd.notna(row[idxs['nw']]) else ""
                            if '/' in raw: parts = raw.split('/'); nw_val = parts[0].strip(); gw_val = parts[1].strip()
                            else: nw_val = raw; gw_val = ""
                        else:
                            nw_val = str(row[idxs['nw']]).strip() if idxs['nw'] != -1 and pd.notna(row[idxs['nw']]) else ""
                            gw_val = str(row[idxs['gw']]).strip() if idxs['gw'] != -1 and pd.notna(row[idxs['gw']]) else ""
                        
                        if nw_val: nw_str = f"N.W: {nw_val}"
                        if gw_val: gw_str = f"G.W: {gw_val}"

                        parts = [f"{str_name} {str_code}", qty_per_str, ctn_count_str, range_str, total_qty_str, nw_str, gw_str]
                        entry = " - ".join([p for p in parts if p])
                        strip_items.append(entry)

                    full_strip = " || \n".join(strip_items)
                    if full_strip:
                        master_data.append({
                            "Week": week_num,
                            "Source File": filename,
                            "Invoice": sheet_name,
                            "POD": pod_val,
                            "Shipment Details": full_strip
                        })
                except: pass
        except: pass

    if master_data:
        log_callback("Saving 'Master_Shipment_DB.xlsx'...")
        output_file = os.path.join(input_folder, "Master_Shipment_DB.xlsx")
        
        saved = False
        while not saved:
            try:
                with open(output_file, 'a'): pass 
                df_out = pd.DataFrame(master_data)
                if 'Week' in df_out.columns: df_out = df_out.sort_values(by=['Week', 'Invoice'])
                df_out.to_excel(output_file, index=False)
                
                wb = load_workbook(output_file)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
                    for cell in row: cell.alignment = Alignment(wrap_text=True, vertical='top')
                ws.column_dimensions['E'].width = 130
                wb.save(output_file)
                saved = True
            except PermissionError:
                if not messagebox.askretrycancel("File Open", "Close 'Master_Shipment_DB.xlsx' and click Retry."): return
            except: saved = True; break

        log_callback("SUCCESS!")
        finish_callback(output_file)
    else:
        log_callback("No data found.")

# ==============================================================================
# PART 2: PACKING LIST GENERATOR LOGIC
# ==============================================================================

def parse_strip_to_rows(details_text):
    """
    Parses the consolidated text string BACK into columns for the form.
    Format: Name (Code) - QTY: X pcs - Y cartons [Range] - Total pcs - N.W - G.W
    """
    lines = details_text.split(" || \n")
    parsed_rows = []
    
    for line in lines:
        line = line.strip()
        if not line: continue
        
        # Extract fields using simple string splitting/regex
        # 1. Weights
        nw, gw = "", ""
        nw_match = re.search(r'N\.W: ([^\-]+)', line)
        if nw_match: nw = nw_match.group(1).strip()
        gw_match = re.search(r'G\.W: ([^\-]+)', line)
        if gw_match: gw = gw_match.group(1).strip()
        
        # 2. Qty
        qty = ""
        # Look for the last number before "pcs" minus "QTY:" prefix
        # This is tricky, so we use the formatted flag we created
        # We look for the bare number followed by pcs, usually appearing as "- 156,000 pcs -"
        # Regex: Dash, space, digits/commas, space, pcs
        total_qty_match = re.search(r'-\s*([\d,]+)\s*pcs\s*(?:-|$)', line)
        if total_qty_match:
            qty = total_qty_match.group(1)
        
        # 3. Cartons
        ctns = ""
        ctn_match = re.search(r'-\s*([\d,]+)\s*cartons', line)
        if ctn_match: ctns = ctn_match.group(1)
        
        # 4. Range
        c_range = ""
        range_match = re.search(r'\[Ctn: ([^\]]+)\]', line)
        if range_match: c_range = range_match.group(1)

        # 5. Name (Everything before the first " - ")
        name = line.split(" - ")[0]
        
        parsed_rows.append([name, qty, ctns, c_range, nw, gw])
        
    return parsed_rows

def generate_packing_list_excel(folder, invoice_data):
    """Creates the formatted Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Packing List"
    
    # Styles
    bold_font = Font(bold=True)
    center_align = Alignment(horizontal='center', vertical='center')
    border_style = Side(border_style="thin")
    border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
    
    # Header Info
    ws['A1'] = "PACKING LIST"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    ws['A1'].alignment = center_align
    
    ws['A3'] = f"Invoice No: {invoice_data['Invoice']}"
    ws['A4'] = f"Destination: {invoice_data['POD']}"
    ws['A5'] = f"Week: {invoice_data['Week']}"
    ws['A3'].font = bold_font
    
    # Table Headers
    headers = ["Item Description", "Total Qty (pcs)", "Cartons", "Ctn Range", "N.W (kg)", "G.W (kg)"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=7, column=col_num, value=header)
        cell.font = bold_font
        cell.alignment = center_align
        cell.border = border
        cell.fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")

    # Data Rows
    raw_text = str(invoice_data['Shipment Details'])
    parsed_rows = parse_strip_to_rows(raw_text)
    
    start_row = 8
    for i, r_data in enumerate(parsed_rows):
        row_num = start_row + i
        for col_num, val in enumerate(r_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.border = border
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            if col_num > 1: cell.alignment = center_align # Center numbers

    # Adjust Widths
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12

    # Save
    safe_inv = str(invoice_data['Invoice']).replace('/', '_')
    out_name = f"PL_{safe_inv}.xlsx"
    out_path = os.path.join(folder, out_name)
    wb.save(out_path)
    return out_path

# ==============================================================================
# PART 3: GUI INTERFACE (TABS)
# ==============================================================================

class ShippingManagerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Shipping Manager Pro")
        self.root.geometry("700x550")
        
        # Create Tabs
        self.notebook = ttk.Notebook(root)
        self.notebook.pack(fill='both', expand=True, padx=10, pady=10)
        
        self.tab_scan = tk.Frame(self.notebook)
        self.tab_gen = tk.Frame(self.notebook)
        
        self.notebook.add(self.tab_scan, text="  1. Scanner (Consolidate)  ")
        self.notebook.add(self.tab_gen, text="  2. Form Generator  ")
        
        self.setup_scan_tab()
        self.setup_gen_tab()
        
        self.working_folder = ""
        self.db_df = None # To store loaded database for generator

    # --- TAB 1: SCANNER ---
    def setup_scan_tab(self):
        frame = tk.Frame(self.tab_scan)
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(frame, text="Step 1: Select Weekly Files Folder", font=("Arial", 10, "bold")).pack(anchor='w')
        
        btn_frame = tk.Frame(frame)
        btn_frame.pack(fill='x', pady=5)
        self.btn_browse = tk.Button(btn_frame, text="Select Folder", command=self.browse_folder)
        self.btn_browse.pack(side='left')
        self.lbl_folder = tk.Label(btn_frame, text="No folder selected", fg="gray")
        self.lbl_folder.pack(side='left', padx=10)
        
        self.progress = ttk.Progressbar(frame, orient='horizontal', mode='determinate')
        self.progress.pack(fill='x', pady=15)
        
        self.log_text = scrolledtext.ScrolledText(frame, height=10)
        self.log_text.pack(fill='both', expand=True)
        
        self.btn_run = tk.Button(frame, text="RUN CONSOLIDATION", command=self.run_scanner, bg="green", fg="white", font=("Arial", 11, "bold"), height=2)
        self.btn_run.pack(fill='x', pady=10)

    # --- TAB 2: GENERATOR ---
    def setup_gen_tab(self):
        frame = tk.Frame(self.tab_gen)
        frame.pack(fill='both', expand=True, padx=20, pady=20)
        
        tk.Label(frame, text="Step 2: Create Packing List from Database", font=("Arial", 10, "bold")).pack(anchor='w')
        
        # Load Button
        tk.Button(frame, text="Refresh/Load Database", command=self.load_db_for_gen).pack(anchor='w', pady=5)
        self.lbl_db_status = tk.Label(frame, text="Database not loaded.", fg="red")
        self.lbl_db_status.pack(anchor='w', padx=5)
        
        tk.Label(frame, text="Select Invoice to Print:", font=("Arial", 9)).pack(anchor='w', pady=(15,0))
        
        # Treeview for selection
        cols = ("Week", "Invoice", "POD")
        self.tree = ttk.Treeview(frame, columns=cols, show='headings', selectmode='browse')
        self.tree.heading("Week", text="Week"); self.tree.column("Week", width=50)
        self.tree.heading("Invoice", text="Invoice"); self.tree.column("Invoice", width=150)
        self.tree.heading("POD", text="POD"); self.tree.column("POD", width=200)
        
        vsb = ttk.Scrollbar(frame, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=vsb.set)
        
        self.tree.pack(side='left', fill='both', expand=True, pady=5)
        vsb.pack(side='right', fill='y', pady=5)
        
        # Generate Button
        btn_gen = tk.Button(self.tab_gen, text="GENERATE FORM", command=self.generate_form, bg="#2196F3", fg="white", font=("Arial", 11, "bold"), height=2)
        btn_gen.pack(fill='x', padx=20, pady=20)

    # --- LOGIC ---
    def browse_folder(self):
        folder = filedialog.askdirectory()
        if folder:
            self.working_folder = folder
            self.lbl_folder.config(text=folder)
            self.log(f"Folder selected: {folder}")
            # Try auto-load DB if exists
            self.load_db_for_gen()

    def log(self, msg):
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)

    def run_scanner(self):
        if not self.working_folder: return messagebox.showerror("Error", "Select folder first.")
        self.btn_run.config(state='disabled')
        threading.Thread(target=process_files, args=(self.working_folder, self.log, self.update_prog, self.on_scan_done)).start()

    def update_prog(self, val):
        self.root.after(0, lambda: self.progress.configure(value=val))

    def on_scan_done(self, output_file):
        self.root.after(0, lambda: self._scan_done_ui(output_file))

    def _scan_done_ui(self, output_file):
        self.btn_run.config(state='normal')
        self.open_file(output_file)
        self.load_db_for_gen() # Auto refresh tab 2

    def load_db_for_gen(self):
        if not self.working_folder: return
        db_path = os.path.join(self.working_folder, "Master_Shipment_DB.xlsx")
        if os.path.exists(db_path):
            try:
                self.db_df = pd.read_excel(db_path)
                # Populate Tree
                self.tree.delete(*self.tree.get_children())
                for idx, row in self.db_df.iterrows():
                    self.tree.insert("", "end", iid=idx, values=(row['Week'], row['Invoice'], row.get('POD', '')))
                self.lbl_db_status.config(text="Database Loaded Successfully", fg="green")
            except Exception as e:
                self.lbl_db_status.config(text=f"Error loading DB: {e}", fg="red")
        else:
            self.lbl_db_status.config(text="Master_Shipment_DB.xlsx not found. Run Scanner first.", fg="red")

    def generate_form(self):
        sel = self.tree.selection()
        if not sel: return messagebox.showwarning("Select Invoice", "Please select a row first.")
        
        idx = int(sel[0])
        row_data = self.db_df.iloc[idx]
        
        try:
            out_path = generate_packing_list_excel(self.working_folder, row_data)
            self.open_file(out_path)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def open_file(self, path):
        try:
            if platform.system() == 'Darwin': subprocess.call(('open', path))
            elif platform.system() == 'Windows': os.startfile(path)
            else: subprocess.call(('xdg-open', path))
        except: pass

if __name__ == "__main__":
    root = tk.Tk()
    app = ShippingManagerApp(root)
    root.mainloop()