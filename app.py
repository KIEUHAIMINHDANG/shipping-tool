import streamlit as st
import pandas as pd
import re
import io
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill

# --- AUTHENTICATION (Simple Password Gate) ---
def check_password():
    """Returns `True` if the user had the correct password."""
    def password_entered():
        if st.session_state["password"] == st.secrets["password"]:
            st.session_state["password_correct"] = True
            del st.session_state["password"]  # Don't store password
        else:
            st.session_state["password_correct"] = False

    if "password_correct" not in st.session_state:
        # First run, show input
        st.text_input(
            "Enter Password", type="password", on_change=password_entered, key="password"
        )
        return False
    elif not st.session_state["password_correct"]:
        # Password incorrect, show input again
        st.text_input(
            "Enter Password", type="password", on_change=password_entered, key="password"
        )
        st.error("😕 Password incorrect")
        return False
    else:
        # Password correct
        return True

# --- LOGIC: NORMALIZATION & HEADERS ---
def normalize_text(text):
    if pd.isna(text): return ""
    text = str(text).lower()
    text = re.sub(r'[\n\r\t]', ' ', text)
    text = re.sub(r'\s+', ' ', text)
    return text.strip()

def find_headers_robust(df):
    limit = min(30, len(df))
    for r in range(limit):
        row_clean = [normalize_text(val) for val in df.iloc[r]]
        idxs = {'name': -1, 'code': -1, 'qty_per_ctn': -1, 'range': -1, 'ctn_count': -1, 'total_qty': -1, 'nw': -1, 'gw': -1}
        for i, val in enumerate(row_clean):
            if "tên sản phẩm" in val or "item description" in val: idxs['name'] = i
            if "mã số" in val or "item code" in val: idxs['code'] = i
            if "1 thùng" in val or "pcs/ctn" in val: idxs['qty_per_ctn'] = i
            if "dải số thùng" in val or "carton no" in val: idxs['range'] = i
            if "số lượng carton" in val or "carton qty" in val: idxs['ctn_count'] = i
            if "lượng xuất" in val or "total qty" in val: idxs['total_qty'] = i
            if "total n.w" in val or ("n.w" in val and idxs['nw'] == -1): idxs['nw'] = i
            if "total g.w" in val or ("g.w" in val and idxs['gw'] == -1): idxs['gw'] = i

        if idxs['name'] != -1 and idxs['total_qty'] != -1:
            return r, idxs
    return None, None

# --- LOGIC: PARSING STRIP ---
def parse_strip_to_rows(details_text):
    lines = details_text.split(" || \n")
    parsed_rows = []
    for line in lines:
        line = line.strip()
        if not line: continue
        nw, gw = "", ""
        nw_match = re.search(r'N\.W: ([^\-]+)', line)
        if nw_match: nw = nw_match.group(1).strip()
        gw_match = re.search(r'G\.W: ([^\-]+)', line)
        if gw_match: gw = gw_match.group(1).strip()
        
        qty = ""
        total_qty_match = re.search(r'-\s*([\d,]+)\s*pcs\s*(?:-|$)', line)
        if total_qty_match: qty = total_qty_match.group(1)
        
        ctns = ""
        ctn_match = re.search(r'-\s*([\d,]+)\s*cartons', line)
        if ctn_match: ctns = ctn_match.group(1)
        
        c_range = ""
        range_match = re.search(r'\[Ctn: ([^\]]+)\]', line)
        if range_match: c_range = range_match.group(1)
        
        name = line.split(" - ")[0]
        parsed_rows.append([name, qty, ctns, c_range, nw, gw])
    return parsed_rows

# --- MAIN APP ---
st.set_page_config(page_title="Shipping Manager Pro", layout="wide")

if check_password():
    st.title("🚢 Shipping Manager Pro (Web Edition)")
    
    tab1, tab2 = st.tabs(["1. Scanner (Consolidate)", "2. Form Generator"])
    
    # === TAB 1: SCANNER ===
    with tab1:
        st.header("Step 1: Upload Weekly Files")
        uploaded_files = st.file_uploader("Drag and drop 'Tuần *.xlsm' files here", accept_multiple_files=True)
        
        if uploaded_files and st.button("Run Consolidation"):
            master_data = []
            progress_bar = st.progress(0)
            
            for f_idx, uploaded_file in enumerate(uploaded_files):
                filename = uploaded_file.name
                week_match = re.search(r"Tuần\s*(\d+)", filename, re.IGNORECASE)
                week_num = int(week_match.group(1)) if week_match else 0
                
                try:
                    # Load from memory bytes
                    xls = pd.ExcelFile(uploaded_file, engine='openpyxl')
                    for sheet_name in xls.sheet_names:
                        if sheet_name.lower() in ['mail', 'mau_mui_tuan', 'sheet1']: continue
                        
                        df_raw = pd.read_excel(xls, sheet_name=sheet_name, header=None, nrows=100)
                        header_row, idxs = find_headers_robust(df_raw)
                        if header_row is None: continue
                        
                        # POD
                        pod_val = "N/A"
                        if len(df_raw) > 6 and pd.notna(df_raw.iloc[6, 2]):
                             pod_val = str(df_raw.iloc[6, 2]).strip().replace('\n', ' ')

                        strip_items = []
                        for i in range(header_row + 1, len(df_raw)):
                            row = df_raw.iloc[i]
                            
                            # Extraction Logic (Simplified for brevity, same as Desktop)
                            item_name = row[idxs['name']]
                            if pd.isna(item_name): continue
                            str_name = str(item_name).strip().replace('\n', ' ')
                            if any(x in str_name.upper() for x in ["TOTAL", "MÃ CÂN", "LÁI XE"]): break
                            
                            str_code = f"({str(row[idxs['code']]).strip()})" if idxs['code'] != -1 and pd.notna(row[idxs['code']]) else ""
                            
                            qty_per_str = ""
                            if idxs['qty_per_ctn'] != -1 and pd.notna(row[idxs['qty_per_ctn']]):
                                qty_per_str = f"QTY: {int(row[idxs['qty_per_ctn']]):,} pcs"
                                
                            ctn_count_str = ""
                            if idxs['ctn_count'] != -1 and pd.notna(row[idxs['ctn_count']]):
                                ctn_count_str = f"{int(row[idxs['ctn_count']]):,} cartons"
                                
                            range_str = ""
                            if idxs['range'] != -1:
                                try:
                                    c1 = str(row[idxs['range']]).strip().replace('.0', '') if pd.notna(row[idxs['range']]) else ""
                                    c2 = str(row[idxs['range']+1]).strip() if pd.notna(row[idxs['range']+1]) else ""
                                    c3 = str(row[idxs['range']+2]).strip().replace('.0', '') if pd.notna(row[idxs['range']+2]) else ""
                                    if c1 or c3: range_str = f"[Ctn: {c1}{c2}{c3}]"
                                except: pass
                                
                            total_qty_str = ""
                            try:
                                val = row[idxs['total_qty']]
                                if isinstance(val, str): val = float(val.replace(',', '').replace('.', ''))
                                if val > 0: total_qty_str = f"{int(val):,} pcs"
                            except: pass
                            
                            nw_str, gw_str = "", ""
                            # Quick logic for weights
                            if idxs['nw'] != -1:
                                val = row[idxs['nw']]
                                if pd.notna(val): 
                                    s_val = str(val)
                                    if '/' in s_val and idxs['nw'] == idxs['gw']:
                                        nw_str = f"N.W: {s_val.split('/')[0]}"
                                        gw_str = f"G.W: {s_val.split('/')[1]}"
                                    else:
                                        nw_str = f"N.W: {val}"

                            if not gw_str and idxs['gw'] != -1:
                                val = row[idxs['gw']]
                                if pd.notna(val): gw_str = f"G.W: {val}"

                            parts = [f"{str_name} {str_code}", qty_per_str, ctn_count_str, range_str, total_qty_str, nw_str, gw_str]
                            strip_items.append(" - ".join([p for p in parts if p]))

                        full_strip = " || \n".join(strip_items)
                        if full_strip:
                            master_data.append({
                                "Week": week_num, "Source File": filename,
                                "Invoice": sheet_name, "POD": pod_val,
                                "Shipment Details": full_strip
                            })
                            
                except Exception as e:
                    st.error(f"Error in {filename}: {e}")
                
                progress_bar.progress((f_idx + 1) / len(uploaded_files))
            
            if master_data:
                df_out = pd.DataFrame(master_data)
                if 'Week' in df_out.columns: df_out = df_out.sort_values(by=['Week', 'Invoice'])
                
                # Convert to Bytes for Download
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df_out.to_excel(writer, index=False)
                    # Simple formatting
                    ws = writer.sheets['Sheet1']
                    ws.column_dimensions['E'].width = 100
                
                st.success(f"Processed {len(master_data)} invoices!")
                st.download_button(
                    label="📥 Download Master Database",
                    data=output.getvalue(),
                    file_name="Master_Shipment_DB.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
                
                # Store in session state for Tab 2
                st.session_state['master_df'] = df_out
            else:
                st.warning("No data found.")

    # === TAB 2: GENERATOR ===
    with tab2:
        st.header("Step 2: Generate Packing List")
        
        # Option to upload if not processed in this session
        uploaded_db = st.file_uploader("Upload Master Database (if not just processed)", type=['xlsx'])
        
        df_db = None
        if 'master_df' in st.session_state:
            df_db = st.session_state['master_df']
            st.info("Using data processed in Tab 1.")
        elif uploaded_db:
            df_db = pd.read_excel(uploaded_db)
            
        if df_db is not None:
            # Dropdown for Invoice
            invoice_list = df_db.apply(lambda x: f"Week {x['Week']} - {x['Invoice']} (POD: {x.get('POD','')})", axis=1).tolist()
            selected_inv_str = st.selectbox("Select Invoice", invoice_list)
            
            if st.button("Generate Form"):
                # Find the row index
                idx = invoice_list.index(selected_inv_str)
                row_data = df_db.iloc[idx]
                
                # Create Excel in Memory
                wb = Workbook()
                ws = wb.active
                ws.title = "Packing List"
                bold_font = Font(bold=True)
                center_align = Alignment(horizontal='center', vertical='center')
                border_style = Side(border_style="thin")
                border = Border(left=border_style, right=border_style, top=border_style, bottom=border_style)
                
                ws['A1'] = "PACKING LIST"; ws['A1'].font = Font(bold=True, size=16); ws.merge_cells('A1:F1'); ws['A1'].alignment = center_align
                ws['A3'] = f"Invoice No: {row_data['Invoice']}"; ws['A3'].font = bold_font
                ws['A4'] = f"Destination: {row_data['POD']}"
                ws['A5'] = f"Week: {row_data['Week']}"
                
                headers = ["Item Description", "Total Qty (pcs)", "Cartons", "Ctn Range", "N.W (kg)", "G.W (kg)"]
                for c_num, h in enumerate(headers, 1):
                    cell = ws.cell(row=7, column=c_num, value=h)
                    cell.font = bold_font; cell.alignment = center_align; cell.border = border; cell.fill = PatternFill(start_color="DDDDDD", fill_type="solid")
                
                raw_text = str(row_data['Shipment Details'])
                parsed_rows = parse_strip_to_rows(raw_text)
                
                start_row = 8
                for i, r_data in enumerate(parsed_rows):
                    for c_num, val in enumerate(r_data, 1):
                        cell = ws.cell(row=start_row+i, column=c_num, value=val)
                        cell.border = border; cell.alignment = Alignment(vertical='top', wrap_text=True)
                        if c_num > 1: cell.alignment = center_align
                
                dims = {'A':50, 'B':15, 'C':12, 'D':15, 'E':12, 'F':12}
                for col, width in dims.items(): ws.column_dimensions[col].width = width
                
                # Save to IO
                output_pl = io.BytesIO()
                wb.save(output_pl)
                safe_inv = str(row_data['Invoice']).replace('/', '_')
                
                st.download_button(
                    label=f"📥 Download PL_{safe_inv}.xlsx",
                    data=output_pl.getvalue(),
                    file_name=f"PL_{safe_inv}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )